"""Tests for long section extraction."""

from __future__ import annotations

import tempfile
from pathlib import Path

import openpyxl
import pytest

from mhls.excel_writer import write_workbook
from mhls.extract_long_sections import (
    _cluster_x_positions,
    _derive_sheet_name,
    _extract_page_rows,
    _get_band_numerics,
    _looks_like_level,
    _snap_to_chainage_columns,
)
from mhls.models import LongSectionRow, LongSectionTable
from tests.fixtures import LONG_SECTION_EXPECTED, LONG_SECTION_WORDS


class TestLooksLikeLevel:
    def test_valid_levels(self) -> None:
        assert _looks_like_level("99.500")
        assert _looks_like_level("0.000")
        assert _looks_like_level("100.123")
        assert _looks_like_level("-5.000")

    def test_invalid(self) -> None:
        assert not _looks_like_level("CHAINAGE")
        assert not _looks_like_level("1200")
        assert not _looks_like_level("abc")
        assert not _looks_like_level("")


class TestGetBandNumerics:
    def test_collects_within_band(self) -> None:
        words = [
            {"text": "0.000", "x0": 100, "x1": 130, "top": 46, "bottom": 54},
            {"text": "5.000", "x0": 200, "x1": 230, "top": 46, "bottom": 54},
            {"text": "OUTSIDE", "x0": 10, "x1": 70, "top": 5, "bottom": 15},
        ]
        result = _get_band_numerics(words, band_y=50.0, band_half_height=8.0)
        assert len(result) == 2

    def test_excludes_outside_band(self) -> None:
        words = [
            {"text": "99.000", "x0": 100, "x1": 130, "top": 200, "bottom": 210},
        ]
        result = _get_band_numerics(words, band_y=50.0, band_half_height=8.0)
        assert len(result) == 0


class TestClusterXPositions:
    def test_groups_nearby_xs(self) -> None:
        words = [
            {"text": "0.000", "x0": 100, "x1": 130, "top": 46, "bottom": 54},
            {"text": "0.000", "x0": 102, "x1": 132, "top": 96, "bottom": 104},
        ]
        clusters = _cluster_x_positions(words, tolerance=10.0)
        assert len(clusters) == 1

    def test_separates_distant_xs(self) -> None:
        words = [
            {"text": "0.000", "x0": 100, "x1": 130, "top": 46, "bottom": 54},
            {"text": "5.000", "x0": 200, "x1": 230, "top": 46, "bottom": 54},
        ]
        clusters = _cluster_x_positions(words, tolerance=10.0)
        assert len(clusters) == 2


class TestSnapToChainageColumns:
    def test_snaps_within_tolerance(self) -> None:
        src_values = {102.0: 99.5, 202.0: 99.2}
        chainage_xs = [100.0, 200.0]
        result = _snap_to_chainage_columns(src_values, chainage_xs, tolerance=12.0)
        assert result[100.0] == pytest.approx(99.5)
        assert result[200.0] == pytest.approx(99.2)

    def test_no_snap_outside_tolerance(self) -> None:
        src_values = {300.0: 99.5}
        chainage_xs = [100.0]
        result = _snap_to_chainage_columns(src_values, chainage_xs, tolerance=12.0)
        assert result[100.0] is None


class TestExtractPageRows:
    def test_all_three_bands_found(self) -> None:
        rows, confidence, found_ch, found_ex, found_al = _extract_page_rows(
            LONG_SECTION_WORDS, "test.pdf", 1
        )
        assert found_ch
        assert found_ex
        assert found_al

    def test_correct_row_count(self) -> None:
        rows, *_ = _extract_page_rows(LONG_SECTION_WORDS, "test.pdf", 1)
        assert len(rows) == len(LONG_SECTION_EXPECTED)

    def test_chainage_values_correct(self) -> None:
        rows, *_ = _extract_page_rows(LONG_SECTION_WORDS, "test.pdf", 1)
        chainages = [r.chainage for r in rows]
        expected_ch = [e[0] for e in LONG_SECTION_EXPECTED]
        assert chainages == pytest.approx(expected_ch)

    def test_existing_level_values_correct(self) -> None:
        rows, *_ = _extract_page_rows(LONG_SECTION_WORDS, "test.pdf", 1)
        existing = [r.existing_level for r in rows]
        expected_ex = [e[1] for e in LONG_SECTION_EXPECTED]
        assert existing == pytest.approx(expected_ex)

    def test_proposed_level_values_correct(self) -> None:
        rows, *_ = _extract_page_rows(LONG_SECTION_WORDS, "test.pdf", 1)
        proposed = [r.proposed_level for r in rows]
        expected_al = [e[2] for e in LONG_SECTION_EXPECTED]
        assert proposed == pytest.approx(expected_al)

    def test_chainage_monotonically_increases(self) -> None:
        rows, *_ = _extract_page_rows(LONG_SECTION_WORDS, "test.pdf", 1)
        chainages = [r.chainage for r in rows]
        for i in range(len(chainages) - 1):
            assert chainages[i] < chainages[i + 1]

    def test_confidence_is_high(self) -> None:
        _, confidence, *_ = _extract_page_rows(LONG_SECTION_WORDS, "test.pdf", 1)
        assert confidence >= 0.5

    def test_no_chainage_band_returns_empty(self) -> None:
        rows, confidence, found_ch, *_ = _extract_page_rows([], "test.pdf", 1)
        assert rows == []
        assert not found_ch
        assert confidence == 0.0


class TestDeriveSheetName:
    def test_rd1_filename(self) -> None:
        name = _derive_sheet_name(Path("7943_005-01 Long section - Rd1.pdf"))
        assert "Rd1" in name or "Road1" in name

    def test_rd2_filename(self) -> None:
        name = _derive_sheet_name(Path("7943_005-02 Long section - Rd2.pdf"))
        assert "Rd2" in name or "Road2" in name

    def test_pos_filename(self) -> None:
        name = _derive_sheet_name(Path("7943_005-03 Long section - POS.pdf"))
        assert "POS" in name

    def test_unknown_filename_uses_stem(self) -> None:
        name = _derive_sheet_name(Path("random_file.pdf"))
        assert name.startswith("LS_")
        assert len(name) <= 31

    def test_max_length_31(self) -> None:
        name = _derive_sheet_name(Path("a" * 100 + ".pdf"))
        assert len(name) <= 31


class TestExcelWriterLongSection:
    """Test that the Excel output has correct sheet structure and blank rows."""

    def _make_ls_table(self, sheet_name: str) -> LongSectionTable:
        rows = [
            LongSectionRow(chainage=0.0, existing_level=99.5, proposed_level=98.0),
            LongSectionRow(chainage=5.0, existing_level=99.2, proposed_level=97.5),
            LongSectionRow(chainage=10.0, existing_level=98.9, proposed_level=97.0),
        ]
        return LongSectionTable(sheet_name=sheet_name, rows=rows)

    def test_sheet_names_correct(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "out.xlsx"
            table = self._make_ls_table("LS_Rd1_Road1")
            write_workbook(out, [], [], [table])

            wb = openpyxl.load_workbook(out)
            assert "MH_Schedule" in wb.sheetnames
            assert "Lateral_MH_Schedule" in wb.sheetnames
            assert "LS_Rd1_Road1" in wb.sheetnames

    def test_ls_headers_correct(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "out.xlsx"
            table = self._make_ls_table("LS_Rd1_Road1")
            write_workbook(out, [], [], [table])

            wb = openpyxl.load_workbook(out)
            ws = wb["LS_Rd1_Road1"]
            assert ws.cell(1, 1).value == "Chainage"
            assert ws.cell(1, 2).value == "Existing Level"
            assert ws.cell(1, 3).value == "Proposed Level"

    def test_blank_row_between_entries(self) -> None:
        """Each data entry must be preceded by a blank row."""
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "out.xlsx"
            table = self._make_ls_table("LS_Rd1_Road1")
            write_workbook(out, [], [], [table])

            wb = openpyxl.load_workbook(out)
            ws = wb["LS_Rd1_Road1"]

            # Row 1 = headers
            # Row 2 = blank, Row 3 = first data
            # Row 4 = blank, Row 5 = second data
            # Row 6 = blank, Row 7 = third data
            assert ws.cell(2, 1).value is None, "Row 2 should be blank"
            assert ws.cell(3, 1).value == pytest.approx(0.0), "Row 3 first chainage"
            assert ws.cell(4, 1).value is None, "Row 4 should be blank"
            assert ws.cell(5, 1).value == pytest.approx(5.0), "Row 5 second chainage"
            assert ws.cell(6, 1).value is None, "Row 6 should be blank"
            assert ws.cell(7, 1).value == pytest.approx(10.0), "Row 7 third chainage"

    def test_numeric_types_in_cells(self) -> None:
        """Values in LS sheet must be stored as numbers, not strings."""
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "out.xlsx"
            table = self._make_ls_table("LS_Rd1_Road1")
            write_workbook(out, [], [], [table])

            wb = openpyxl.load_workbook(out)
            ws = wb["LS_Rd1_Road1"]
            # Row 3 is first data row
            assert isinstance(ws.cell(3, 1).value, (int, float))
            assert isinstance(ws.cell(3, 2).value, (int, float))
            assert isinstance(ws.cell(3, 3).value, (int, float))


class TestExcelWriterMHSchedule:
    def test_mh_schedule_headers(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "out.xlsx"
            write_workbook(out, [], [], [])

            wb = openpyxl.load_workbook(out)
            ws = wb["MH_Schedule"]
            assert ws.cell(1, 1).value == "MH REF"
            assert ws.cell(1, 2).value == "MH DIA (size)"
            assert ws.cell(1, 3).value == "Cover Level"
            assert ws.cell(1, 4).value == "Invert Level"

    def test_lateral_schedule_headers(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "out.xlsx"
            write_workbook(out, [], [], [])

            wb = openpyxl.load_workbook(out)
            ws = wb["Lateral_MH_Schedule"]
            assert ws.cell(1, 1).value == "MH REF"
            assert ws.cell(1, 5).value == "Diameter"

    def test_mh_rows_no_blank_rows(self) -> None:
        """MH schedule rows must have no blank rows between them."""
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "out.xlsx"
            from mhls.models import MHScheduleRow

            rows = [
                MHScheduleRow("S1", 1200, 50.25, 47.05),
                MHScheduleRow("S2", 1200, 50.10, 46.90),
            ]
            write_workbook(out, rows, [], [])

            wb = openpyxl.load_workbook(out)
            ws = wb["MH_Schedule"]
            # Row 1 = header, Row 2 = S1, Row 3 = S2 (no blanks)
            assert ws.cell(2, 1).value == "S1"
            assert ws.cell(3, 1).value == "S2"
            assert ws.cell(4, 1).value is None
