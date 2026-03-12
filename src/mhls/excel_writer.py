"""Write extracted data to an Excel workbook using openpyxl."""

from __future__ import annotations

import logging
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

from mhls.models import LateralMHScheduleRow, LongSectionTable, MHScheduleRow

log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Sheet headers
# ---------------------------------------------------------------------------

_MH_HEADERS = ["MH REF", "MH DIA (size)", "Cover Level", "Invert Level"]
_LATERAL_HEADERS = ["MH REF", "MH DIA (size)", "Cover Level", "Invert Level", "Diameter"]
_LS_HEADERS = ["Chainage", "Existing Level", "Proposed Level"]


def _write_header(ws, headers: list[str]) -> None:
    """Write bold header row to worksheet."""
    for col, heading in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=heading)
        cell.font = Font(bold=True)


# ---------------------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------------------


def _write_mh_schedule_sheet(
    wb: openpyxl.Workbook,
    rows: list[MHScheduleRow],
) -> None:
    """Write MH_Schedule sheet."""
    ws = wb.create_sheet("MH_Schedule")
    _write_header(ws, _MH_HEADERS)
    for row_num, row in enumerate(rows, start=2):
        ws.cell(row=row_num, column=1, value=row.mh_ref)
        ws.cell(row=row_num, column=2, value=row.mh_dia)
        ws.cell(row=row_num, column=3, value=row.cover_level)
        ws.cell(row=row_num, column=4, value=row.invert_level)
    log.info("MH_Schedule sheet written with %d rows", len(rows))


def _write_lateral_mh_schedule_sheet(
    wb: openpyxl.Workbook,
    rows: list[LateralMHScheduleRow],
) -> None:
    """Write Lateral_MH_Schedule sheet."""
    ws = wb.create_sheet("Lateral_MH_Schedule")
    _write_header(ws, _LATERAL_HEADERS)
    for row_num, row in enumerate(rows, start=2):
        ws.cell(row=row_num, column=1, value=row.mh_ref)
        ws.cell(row=row_num, column=2, value=row.mh_dia)
        ws.cell(row=row_num, column=3, value=row.cover_level)
        ws.cell(row=row_num, column=4, value=row.invert_level)
        ws.cell(row=row_num, column=5, value=row.diameter)
    log.info("Lateral_MH_Schedule sheet written with %d rows", len(rows))


def _write_long_section_sheet(
    wb: openpyxl.Workbook,
    table: LongSectionTable,
) -> None:
    """Write a long section sheet with a blank row between each entry.

    Layout:
      Row 1: headers
      Row 2: blank
      Row 3: first data row
      Row 4: blank
      Row 5: second data row
      ...
    """
    ws = wb.create_sheet(table.sheet_name)
    _write_header(ws, _LS_HEADERS)

    # Write blank row then data row for each entry
    current_row = 2
    for ls_row in table.rows:
        # Blank row
        # (leave current_row empty - just increment)
        current_row += 1
        # Data row
        ws.cell(row=current_row, column=1, value=ls_row.chainage)
        ws.cell(row=current_row, column=2, value=ls_row.existing_level)
        ws.cell(row=current_row, column=3, value=ls_row.proposed_level)
        current_row += 1

    log.info("Sheet '%s' written with %d data rows", table.sheet_name, len(table.rows))


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def write_workbook(
    output_path: Path,
    mh_rows: list[MHScheduleRow],
    lateral_rows: list[LateralMHScheduleRow],
    long_section_tables: list[LongSectionTable],
) -> None:
    """Write all data to a single Excel workbook."""
    wb = openpyxl.Workbook()

    # Remove the default empty sheet
    default_sheet = wb.active
    if default_sheet is not None:
        wb.remove(default_sheet)

    _write_mh_schedule_sheet(wb, mh_rows)
    _write_lateral_mh_schedule_sheet(wb, lateral_rows)

    for table in long_section_tables:
        _write_long_section_sheet(wb, table)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    log.info("Workbook written to %s", output_path)
